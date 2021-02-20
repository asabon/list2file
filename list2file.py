import sys
from openpyxl import Workbook
from openpyxl import load_workbook

def list2file(listfile, fileindex, valueindex, target_row, target_col):
    wb = load_workbook(filename=listfile)
    ws = wb.active
    for i in range (2, 10):
        targetfile = ws.cell(row = i, column = int(fileindex)).value
        value      = ws.cell(row = i, column = int(valueindex)).value
        if (targetfile != None):
            wb2 = load_workbook(filename=targetfile)
            ws2 = wb2.active
            ws2.cell(row = int(target_row), column = int(target_col)).value = value
            wb2.save(targetfile)
            wb2.close()
    wb.close()

#####################################################################################
#
# Usage
#    python list2file.py listfile fileindex valueindex target_row target_col
#
#    listfile   : list file name
#    fileindex  : filename column index in listfile
#    valueindex : value column index in listfile
#    target_row : row number to write value to each file 
#    target_col : row number to write value to each file
#
#####################################################################################
if __name__ == '__main__':
    value = sys.argv
    if len(value) == 6:
        listfile = value[1]
        fileindex = value[2]
        valueindex = value[3]
        target_row = value[4]
        target_col = value[5]
        list2file(listfile, fileindex, valueindex, target_row, target_col)
    else:
        print ("argument error")
